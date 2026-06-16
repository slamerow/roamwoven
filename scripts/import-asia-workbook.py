import json
from pathlib import Path

import openpyxl


SOURCE = Path("/Users/eli/Downloads/Backend for Asia Trip - Current Working Doc.xlsx")
OUTPUT = Path("data/asia-trip-seed.json")


def row_dict(headers, row):
    return {
        header: value
        for header, value in zip(headers, row)
        if header is not None and value is not None
    }


def iso(value):
    if hasattr(value, "isoformat"):
        return value.isoformat()
    return value


def main():
    workbook = openpyxl.load_workbook(SOURCE, data_only=True)

    legs_sheet = workbook["Legs"]
    activities_sheet = workbook["Activities"]

    leg_headers = [cell.value for cell in legs_sheet[1]]
    activity_headers = [cell.value for cell in activities_sheet[1]]

    legs = []
    for row in legs_sheet.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        leg = row_dict(leg_headers, row)
        if "leg_id" not in leg:
            continue
        legs.append(
            {
                "id": str(leg["leg_id"]),
                "country": leg.get("country"),
                "city": leg.get("city"),
                "arriveDate": iso(leg.get("arrive")),
                "leaveDate": iso(leg.get("leave")),
                "nights": leg.get("nights"),
                "stayName": leg.get("stay_name"),
                "stayAddress": leg.get("stay_address"),
                "why": leg.get("why"),
                "timezone": leg.get("timezone"),
                "language": leg.get("language"),
                "latitude": leg.get("latitude"),
                "longitude": leg.get("longitude"),
            }
        )

    items = []
    for row in activities_sheet.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        item = row_dict(activity_headers, row)
        if "activity_id" not in item:
            continue
        title = item.get("title")
        if not title:
            continue
        items.append(
            {
                "id": str(item["activity_id"]),
                "legId": str(item.get("leg_id")) if item.get("leg_id") else None,
                "date": iso(item.get("date")),
                "startTime": iso(item.get("start_time")),
                "endTime": iso(item.get("end_time")),
                "title": title,
                "description": item.get("description"),
                "category": item.get("category"),
                "locationName": item.get("location_name"),
                "address": item.get("address"),
                "url": item.get("url"),
                "notes": item.get("notes"),
            }
        )

    payload = {
        "name": "Wren's Adventure",
        "dateRange": "June 27 - October 10, 2026",
        "source": "Backend for Asia Trip - Current Working Doc.xlsx",
        "legs": legs,
        "items": items,
    }

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT.write_text(json.dumps(payload, indent=2, ensure_ascii=False) + "\n")
    print(f"Wrote {len(legs)} legs and {len(items)} items to {OUTPUT}")


if __name__ == "__main__":
    main()
